"""studio_import_excel.xlsx 模板 (8 sheet) ↔ 图状态 转换层。

本模块是「导出 + 增量回灌」spec 的
核心 I/O 层：

- ``export_to_workbook(driver, datasource_id)``：读图 → 单 workbook（8 sheet +
  隐藏 ``_meta``）。
- ``parse_import_workbook(path)``：用户编辑后的 xlsx → 内部 bundle dict。
- ``build_bundle_from_graph(...)`` / ``_query_*``：纯 Cypher 直读（不依赖
  ``semantic_store`` 的部分裁剪返回，因为导出需要全字段：unit / is_north_star /
  derived_from / evidence_ext / dim binding 全量 / dim value 全量）。

模板 schema（每行首列均为 ``datasource_id``）：

| sheet | bold key 列 | 非多值 / 多值 |
|---|---|---|
| datasource | datasource_id | 单值 |
| biz_domain | datasource_id, domain_name | 单值 |
| dataset | datasource_id, domain_name, dataset_name | 单值 |
| dataset_column | datasource_id, domain_name, dataset_name, column_name | 单值 |
| dimension | datasource_id, domain_name, dimension_name | 单值 |
| dataset_dimension | （无 bold，整行 = key） | 多值：全列拼接 |
| metric | datasource_id, domain_name, metric_name | 单值 |
| metric_formula | （无 bold，整行 = key） | 多值：全列拼接 |

「多值那种 list，回写用 ``$$$`` 拼接」：指 ``synonyms`` / ``aliases`` / ``tags`` /
``enums`` / ``samples`` / ``column_enums`` / ``derived_from`` 这类数组字段，
统一以 ``$$$`` 连接成单格字符串（回灌时反向 split）。
``dataset_dimension`` / ``metric_formula`` 无 bold key 列 → 用整行所有列 ``$$$`` 拼接做复合 key。
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import time
import uuid
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Optional

from neo4j import Driver

from ..utils import get_logger, neo4j_session
from .keys import (
    caliber_key,
    dataset_column_key,
    dataset_key,
    dim_key,
    dim_value_key,
    domain_key,
    formula_key,
    metric_key,
)

log = get_logger("graph.semantic_template_io")

# ---------------------------------------------------------------------- #
# 模板 schema 定义
# ---------------------------------------------------------------------- #

# 每个 sheet 的列顺序（与 studio_import_excel.xlsx 完全一致）
SHEET_COLUMNS: dict[str, list[str]] = {
    "datasource": ["datasource_id", "datasource_name", "datasource_type"],
    "biz_domain": ["datasource_id", "domain_name", "display_name", "description", "aliases"],
    "dataset": ["datasource_id", "domain_name", "dataset_name", "dataset_comment",
                "dataset_type", "sql_content", "parents"],
    "dataset_column": ["datasource_id", "domain_name", "dataset_name", "column_name",
                       "is_primary", "is_nullable", "data_type", "column_comment",
                       "column_enums", "column_enums_description", "column_type",
                       "samples", "dimension_type", "column_name_cn"],
    "dimension": ["datasource_id", "domain_name", "dimension_name", "description",
                  "parent_name", "depth", "synonyms", "is_visible", "is_attribution",
                  "enums"],
    "dataset_dimension": ["datasource_id", "domain_name", "dataset_name",
                          "dimension_name", "calculate_expr", "dimension_type",
                          "data_type"],
    "metric": ["datasource_id", "domain_name", "metric_name", "description", "unit",
               "is_polaris", "show_distribution", "is_visible", "synonyms", "tags"],
    "metric_formula": ["datasource_id", "domain_name", "metric_name", "dataset_name",
                       "formula", "date_range", "formula_evidence", "derived_from",
                       "evidence_ext"],
}

# bold = key 列（用于唯一标识一行）。空列表 = 无 bold，整行做 key。
SHEET_KEY_COLUMNS: dict[str, list[str]] = {
    "datasource": ["datasource_id"],
    "biz_domain": ["datasource_id", "domain_name"],
    "dataset": ["datasource_id", "domain_name", "dataset_name"],
    "dataset_column": ["datasource_id", "domain_name", "dataset_name", "column_name"],
    "dimension": ["datasource_id", "domain_name", "dimension_name"],
    "dataset_dimension": [],  # 无 bold → 整行拼接
    "metric": ["datasource_id", "domain_name", "metric_name"],
    "metric_formula": [],  # 无 bold → 整行拼接
}

# 多值列：这些列在图里是 list，写 cell 时用 $$$ 拼接，读 cell 时反向 split。
MULTIVALUE_COLUMNS: set[str] = {
    "aliases", "synonyms", "tags", "enums", "samples", "column_enums",
    "derived_from",
}

# 列顺序里 sheet 出现的顺序（导出 workbook 时按此建 sheet）
SHEET_ORDER: list[str] = [
    "datasource", "biz_domain", "dataset", "dataset_column",
    "dimension", "dataset_dimension", "metric", "metric_formula",
]

# 布尔列：True→"1"，False→"0"
BOOL_COLUMNS: set[str] = {
    "is_primary", "is_nullable", "is_polaris", "show_distribution",
    "is_visible", "is_attribution",
}

_LIST_SEP = "$$$"


def _join_list(vals: Iterable[Any]) -> str:
    """list → ``$$$`` 拼接的字符串；None/空 → ''。"""
    out: list[str] = []
    for v in vals:
        if v is None:
            continue
        s = str(v).strip()
        if s:
            out.append(s)
    return _LIST_SEP.join(out)


def _split_list(val: Any) -> list[str]:
    """``$$$`` 拼接的字符串 → list；空 → []。兼容旧格式（逗号/顿号/竖线）。"""
    if not val:
        return []
    s = str(val).strip()
    if not s:
        return []
    if _LIST_SEP in s:
        return [p.strip() for p in s.split(_LIST_SEP) if p.strip()]
    # 兼容：无 $$$ 但有其它分隔符
    parts = re.split(r"[,，、;；|]", s)
    return [p.strip() for p in parts if p.strip()]


def _bool_to_cell(v: Any) -> str:
    if v is None:
        return "0"
    if isinstance(v, bool):
        return "1" if v else "0"
    s = str(v).strip().lower()
    return "1" if s in ("1", "true", "yes", "y") else "0"


def _cell_to_bool(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("1", "true", "yes", "y")


def _row_key(sheet: str, row: dict[str, Any]) -> str:
    """计算一行的复合 key（bold 列拼接，或整行拼接）。"""
    key_cols = SHEET_KEY_COLUMNS.get(sheet, [])
    if key_cols:
        parts = [str(row.get(c) or "").strip() for c in key_cols]
    else:
        parts = [str(row.get(c) or "").strip() for c in SHEET_COLUMNS[sheet]]
    return _LIST_SEP.join(parts)


# ---------------------------------------------------------------------- #
# 图直读 Cypher（导出用，取全字段）
# ---------------------------------------------------------------------- #

def _ds_clause(alias: str = "n") -> str:
    return f"($ds = '' OR {alias}.datasource_id = $ds)"


def _query_datasource_row(driver: Driver, ds: str) -> dict[str, Any]:
    """datasource sheet 单行（来自 registry；回退用 ds code 自身）。"""
    from .datasource_registry import try_resolve
    rec = try_resolve(ds)
    name = (rec.display_name if rec else ds) or ds
    ds_type = (rec.db_type if rec else "") or ""
    return {
            "datasource_id": ds,
            "datasource_name": name,
            "datasource_type": ds_type,
    }


def _query_domain_rows(driver: Driver, ds: str) -> list[dict[str, Any]]:
    cypher = f"""
    MATCH (d:Domain)
    WHERE {_ds_clause("d")}
    RETURN d.name AS domain_name,
           coalesce(d.display_name, d.name) AS display_name,
           coalesce(d.description, '') AS description,
           coalesce(d.aliases, []) AS aliases
    ORDER BY domain_name
    """
    with neo4j_session(driver) as s:
        rows = s.run(cypher, ds=ds).data()
    out: list[dict[str, Any]] = []
    for r in rows:
        out.append({
            "datasource_id": ds,
            "domain_name": str(r.get("domain_name") or ""),
            "display_name": str(r.get("display_name") or r.get("domain_name") or ""),
            "description": str(r.get("description") or ""),
            "aliases": _join_list(r.get("aliases") or []),
        })
    return out


def _query_dataset_rows(driver: Driver, ds: str) -> list[dict[str, Any]]:
    cypher = f"""
    MATCH (ds:Dataset)
    WHERE {_ds_clause("ds")}
    RETURN ds.name AS dataset_name,
           coalesce(ds.description, '') AS dataset_comment,
           coalesce(ds.dataset_type, 'OLAP') AS dataset_type,
           coalesce(ds.sql, '') AS sql_content,
           coalesce(ds.parents, '') AS parents,
           ds.domain AS domain_name
    ORDER BY domain_name, dataset_name
    """
    with neo4j_session(driver) as s:
        rows = s.run(cypher, ds=ds).data()
    out: list[dict[str, Any]] = []
    for r in rows:
        out.append({
            "datasource_id": ds,
            "domain_name": str(r.get("domain_name") or ""),
            "dataset_name": str(r.get("dataset_name") or ""),
            "dataset_comment": str(r.get("dataset_comment") or ""),
            "dataset_type": str(r.get("dataset_type") or "OLAP"),
            "sql_content": str(r.get("sql_content") or ""),
            "parents": str(r.get("parents") or ""),
        })
    return out


def _query_dataset_column_rows(driver: Driver, ds: str) -> list[dict[str, Any]]:
    cypher = (
        "MATCH (ds:Dataset)-[:HAS_COLUMN]->(dc:DatasetColumn)\n"
        "WHERE ($ds = '' OR ds.datasource_id = $ds)\n"
        "RETURN ds.domain AS domain_name,\n"
        "       ds.name AS dataset_name,\n"
        "       dc.name AS column_name,\n"
        "       coalesce(dc.is_primary, false) AS is_primary,\n"
        "       coalesce(dc.is_nullable, true) AS is_nullable,\n"
        "       coalesce(dc.data_type, 'text') AS data_type,\n"
        "       coalesce(dc.description, '') AS column_comment,\n"
        "       coalesce(dc.column_enums, []) AS column_enums,\n"
        "       coalesce(dc.enum_value_descriptions, {}) AS column_enums_description,\n"
        "       coalesce(dc.column_type, '') AS column_type,\n"
        "       coalesce(dc.sample_values, []) AS samples,\n"
        "       coalesce(dc.dimension_type, '') AS dimension_type,\n"
        "       coalesce(dc.display_name, dc.name) AS column_name_cn\n"
        "ORDER BY domain_name, dataset_name, column_name"
    )
    with neo4j_session(driver) as s:
        rows = s.run(cypher, ds=ds).data()
    out: list[dict[str, Any]] = []
    for r in rows:
        # column_enums_description 是 dict → JSON 字符串
        ced = r.get("column_enums_description") or {}
        ced_str = json.dumps(ced, ensure_ascii=False) if ced else ""
        out.append({
            "datasource_id": ds,
            "domain_name": str(r.get("domain_name") or ""),
            "dataset_name": str(r.get("dataset_name") or ""),
            "column_name": str(r.get("column_name") or ""),
            "is_primary": _bool_to_cell(r.get("is_primary")),
            "is_nullable": _bool_to_cell(r.get("is_nullable")),
            "data_type": str(r.get("data_type") or "text"),
            "column_comment": str(r.get("column_comment") or ""),
            "column_enums": _join_list(r.get("column_enums") or []),
            "column_enums_description": ced_str,
            "column_type": str(r.get("column_type") or ""),
            "samples": _join_list(r.get("samples") or []),
            "dimension_type": str(r.get("dimension_type") or ""),
            "column_name_cn": str(r.get("column_name_cn") or ""),
        })
    return out


def _query_dimension_rows(driver: Driver, ds: str) -> list[dict[str, Any]]:
    cypher = f"""
    MATCH (d:Dimension)
    WHERE {_ds_clause("d")}
    OPTIONAL MATCH (d)-[:HAS_PARENT]->(p:Dimension)
    OPTIONAL MATCH (d)-[:HAS_VALUE]->(dv:DimensionValue)
    WITH d, p, collect(DISTINCT {'{value: dv.value, occur_cnt: coalesce(dv.occur_cnt, 0)}'}) AS vals
    RETURN d.domain AS domain_name,
           d.name AS dimension_name,
           coalesce(d.description, '') AS description,
           coalesce(p.name, '') AS parent_name,
           coalesce(d.hierarchy_level, 0) AS depth,
           coalesce(d.aliases, []) AS synonyms,
           coalesce(d.is_display_dimension, true) AS is_visible,
           coalesce(d.is_contribution_dimension, true) AS is_attribution,
           vals
    ORDER BY domain_name, dimension_name
    """
    with neo4j_session(driver) as s:
        rows = s.run(cypher, ds=ds).data()
    out: list[dict[str, Any]] = []
    for r in rows:
        vals = r.get("vals") or []
        enums = [
            {"value": v.get("value"), "occur_cnt": v.get("occur_cnt")}
            for v in vals if v and v.get("value") is not None
        ]
        out.append({
            "datasource_id": ds,
            "domain_name": str(r.get("domain_name") or ""),
            "dimension_name": str(r.get("dimension_name") or ""),
            "description": str(r.get("description") or ""),
            "parent_name": str(r.get("parent_name") or ""),
            "depth": str(r.get("depth") or "0"),
            "synonyms": _join_list(r.get("synonyms") or []),
            "is_visible": _bool_to_cell(r.get("is_visible")),
            "is_attribution": _bool_to_cell(r.get("is_attribution")),
            "enums": json.dumps(enums, ensure_ascii=False) if enums else "",
        })
    return out


def _query_dataset_dimension_rows(driver: Driver, ds: str) -> list[dict[str, Any]]:
    cypher = f"""
    MATCH (d:Dimension)-[r:MAPS_TO_DATASET_COLUMN|MAPS_TO_COLUMN]->(target)
    WHERE {_ds_clause("d")}
    OPTIONAL MATCH (target)<-[:HAS_COLUMN]-(ds:Dataset)
    RETURN d.domain AS domain_name,
           d.name AS dimension_name,
           coalesce(ds.name, '') AS dataset_name,
           coalesce(r.expr, '') AS calculate_expr,
           coalesce(d.dimension_type, 'OLAP维度') AS dimension_type,
           coalesce(d.data_type, 'text') AS data_type
    ORDER BY domain_name, dataset_name, dimension_name
    """
    with neo4j_session(driver) as s:
        rows = s.run(cypher, ds=ds).data()
    out: list[dict[str, Any]] = []
    for r in rows:
        expr = str(r.get("calculate_expr") or "")
        if expr and not expr.lower().startswith("select"):
            expr = f"select({expr})"
        out.append({
            "datasource_id": ds,
            "domain_name": str(r.get("domain_name") or ""),
            "dataset_name": str(r.get("dataset_name") or ""),
            "dimension_name": str(r.get("dimension_name") or ""),
            "calculate_expr": expr,
            "dimension_type": str(r.get("dimension_type") or "OLAP维度"),
            "data_type": str(r.get("data_type") or "text"),
        })
    return out


def _query_metric_rows(driver: Driver, ds: str) -> list[dict[str, Any]]:
    cypher = f"""
    MATCH (m:Metric)
    WHERE {_ds_clause("m")}
    RETURN m.domain AS domain_name,
           m.name AS metric_name,
           coalesce(m.description, '') AS description,
           coalesce(m.unit, '') AS unit,
           coalesce(m.is_north_star, false) AS is_polaris,
           coalesce(m.is_display_distribution, false) AS show_distribution,
           coalesce(m.is_display, true) AS is_visible,
           coalesce(m.aliases, []) AS synonyms,
           coalesce(m.tags, []) AS tags
    ORDER BY domain_name, metric_name
    """
    with neo4j_session(driver) as s:
        rows = s.run(cypher, ds=ds).data()
    out: list[dict[str, Any]] = []
    for r in rows:
        out.append({
            "datasource_id": ds,
            "domain_name": str(r.get("domain_name") or ""),
            "metric_name": str(r.get("metric_name") or ""),
            "description": str(r.get("description") or ""),
            "unit": str(r.get("unit") or ""),
            "is_polaris": _bool_to_cell(r.get("is_polaris")),
            "show_distribution": _bool_to_cell(r.get("show_distribution")),
            "is_visible": _bool_to_cell(r.get("is_visible")),
            "synonyms": _join_list(r.get("synonyms") or []),
            "tags": _join_list(r.get("tags") or []),
        })
    return out


def _query_metric_formula_rows(driver: Driver, ds: str) -> list[dict[str, Any]]:
    # Formula + OF_VIEW→Dataset + DERIVED_FROM（图里可能没有 DERIVED_FROM 边，
    # 退化为空）。derived_from 为 JSON list → 直接吐原字符串。
    cypher = f"""
    MATCH (m:Metric)-[:HAS_FORMULA]->(f:Formula)
    WHERE {_ds_clause("m")}
    OPTIONAL MATCH (f)-[:OF_VIEW]->(ds:Dataset)
    OPTIONAL MATCH (m)-[df:DERIVED_FROM]->(peer:Metric)
    WITH m, f, ds, collect(DISTINCT {{name: peer.name, role: coalesce(df.role, '')}}) AS peers
    RETURN m.domain AS domain_name,
           m.name AS metric_name,
           coalesce(ds.name, f.dataset, '') AS dataset_name,
           coalesce(f.formula, '') AS formula,
           coalesce(f.date_range, f.refresh_freq, '') AS date_range,
           coalesce(f.formula_evidence, '') AS formula_evidence,
           peers,
           coalesce(f.evidence_ext, '') AS evidence_ext
    ORDER BY domain_name, metric_name, dataset_name, date_range
    """
    with neo4j_session(driver) as s:
        rows = s.run(cypher, ds=ds).data()
    out: list[dict[str, Any]] = []
    for r in rows:
        peers = r.get("peers") or []
        derived: list[str] = []
        for p in peers:
            n = str(p.get("name") or "").strip()
            if n:
                role = str(p.get("role") or "").strip()
                if role in ("numerator", "denominator"):
                    derived.append(f"{n}:{role}")
                else:
                    derived.append(n)
        out.append({
            "datasource_id": ds,
            "domain_name": str(r.get("domain_name") or ""),
            "metric_name": str(r.get("metric_name") or ""),
            "dataset_name": str(r.get("dataset_name") or ""),
            "formula": str(r.get("formula") or ""),
            "date_range": str(r.get("date_range") or ""),
            "formula_evidence": str(r.get("formula_evidence") or ""),
            "derived_from": _LIST_SEP.join(derived) if derived else "",
            "evidence_ext": str(r.get("evidence_ext") or ""),
        })
    return out


# ---------------------------------------------------------------------- #
# baseline digest / entity keys
# ---------------------------------------------------------------------- #

@dataclass
class ExportBaseline:
    """导出时刻的图快照指纹，供回灌 diff + 漂移守卫使用。"""
    export_id: str
    datasource_id: str
    exported_at: float
    graph_digest: str  # 当前图状态哈希
    # 按 sheet → set[row_key] 的 baseline key 清单
    entity_keys: dict[str, list[str]] = field(default_factory=dict)

    def to_meta_rows(self) -> list[dict[str, Any]]:
        """_meta sheet 的行（扁平化便于 xlsx 存储）。"""
        return [
            {"key": "export_id", "value": self.export_id},
            {"key": "datasource_id", "value": self.datasource_id},
            {"key": "exported_at", "value": str(self.exported_at)},
            {"key": "graph_digest", "value": self.graph_digest},
            {"key": "entity_keys", "value": json.dumps(self.entity_keys, ensure_ascii=False)},
        ]


def _compute_graph_digest(rows_by_sheet: dict[str, list[dict[str, Any]]]) -> str:
    """对所有 sheet 行的 (row_key → 排序后序列化) 算 sha1。"""
    h = hashlib.sha1()
    for sheet in SHEET_ORDER:
        rows = rows_by_sheet.get(sheet, [])
        keys = sorted(_row_key(sheet, r) for r in rows)
        h.update(sheet.encode())
        h.update(b"\x00")
        for k in keys:
            h.update(k.encode())
            h.update(b"\x01")
        h.update(b"\x02")
    return "sha1:" + h.hexdigest()


def _collect_entity_keys(rows_by_sheet: dict[str, list[dict[str, Any]]]) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    for sheet in SHEET_ORDER:
        out[sheet] = sorted({_row_key(sheet, r) for r in rows_by_sheet.get(sheet, [])})
    return out


# ---------------------------------------------------------------------- #
# 导出：图 → workbook
# ---------------------------------------------------------------------- #

def build_bundle_from_graph(driver: Driver, datasource_id: str) -> dict[str, Any]:
    """读图 → 模板 bundle（8 sheet 的行 + meta）。纯读，不写。"""
    ds = (datasource_id or "").strip()
    rows_by_sheet: dict[str, list[dict[str, Any]]] = {
        "datasource": [_query_datasource_row(driver, ds)] if ds else [],
        "biz_domain": _query_domain_rows(driver, ds),
        "dataset": _query_dataset_rows(driver, ds),
        "dataset_column": _query_dataset_column_rows(driver, ds),
        "dimension": _query_dimension_rows(driver, ds),
        "dataset_dimension": _query_dataset_dimension_rows(driver, ds),
        "metric": _query_metric_rows(driver, ds),
        "metric_formula": _query_metric_formula_rows(driver, ds),
    }
    baseline = ExportBaseline(
        export_id=f"exp_{uuid.uuid4().hex[:12]}",
        datasource_id=ds,
        exported_at=time.time(),
        graph_digest=_compute_graph_digest(rows_by_sheet),
        entity_keys=_collect_entity_keys(rows_by_sheet),
    )
    return {"rows_by_sheet": rows_by_sheet, "meta": baseline}


def export_to_workbook(driver: Driver, datasource_id: str) -> tuple[Any, ExportBaseline]:
    """读图 → openpyxl Workbook（8 sheet + 隐藏 _meta）。

    返回 (workbook, baseline)。调用方负责把 workbook 写到文件/响应流。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl required for export") from exc

    bundle = build_bundle_from_graph(driver, datasource_id)
    rows_by_sheet: dict[str, list[dict[str, Any]]] = bundle["rows_by_sheet"]
    meta: ExportBaseline = bundle["meta"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold_font = Font(bold=True)

    for sheet in SHEET_ORDER:
        ws = wb.create_sheet(title=sheet)
        cols = SHEET_COLUMNS[sheet]
        key_cols = SHEET_KEY_COLUMNS.get(sheet, [])
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            if col in key_cols:
                cell.font = bold_font
        for ri, row in enumerate(rows_by_sheet.get(sheet, []), start=2):
            for ci, col in enumerate(cols, start=1):
                ws.cell(row=ri, column=ci, value=row.get(col, ""))

    # _meta（隐藏）
    ws_meta = wb.create_sheet(title="_meta")
    ws_meta.cell(row=1, column=1, value="key").font = bold_font
    ws_meta.cell(row=1, column=2, value="value").font = bold_font
    for ri, r in enumerate(meta.to_meta_rows(), start=2):
        ws_meta.cell(row=ri, column=1, value=r["key"])
        ws_meta.cell(row=ri, column=2, value=r["value"])
    ws_meta.sheet_state = "hidden"

    return wb, meta


def save_workbook(wb, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ---------------------------------------------------------------------- #
# 回灌：xlsx → 内部 bundle（供 diff 使用）
# ---------------------------------------------------------------------- #

@dataclass
class ParsedImport:
    """用户编辑后 xlsx 的解析结果。"""
    datasource_id: str
    rows_by_sheet: dict[str, list[dict[str, Any]]]
    meta: Optional[ExportBaseline]  # 可能缺失（用户删了 _meta 或非导出文件）
    file_keys: dict[str, list[str]]  # 按 sheet → sorted row_key（F）


def _positive_env_int(name: str, default: int) -> int:
    try:
        return max(1, int(os.getenv(name, str(default))))
    except ValueError:
        return default


def _validate_xlsx_archive(path: Path) -> None:
    """在 openpyxl 解压前限制 ZIP 条目、展开大小和压缩比。"""
    max_entries = _positive_env_int("DATAPAW_XLSX_MAX_ENTRIES", 1000)
    max_uncompressed = (
        _positive_env_int("DATAPAW_XLSX_MAX_UNCOMPRESSED_MB", 200)
        * 1024
        * 1024
    )
    max_ratio = _positive_env_int("DATAPAW_XLSX_MAX_COMPRESSION_RATIO", 100)
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            if len(infos) > max_entries:
                raise ValueError(
                    f"xlsx archive has too many entries ({len(infos)} > {max_entries})"
                )
            if any(info.flag_bits & 0x1 for info in infos):
                raise ValueError("encrypted xlsx archives are not supported")
            unpacked = sum(info.file_size for info in infos)
            packed = sum(info.compress_size for info in infos)
            if unpacked > max_uncompressed:
                raise ValueError(
                    "xlsx uncompressed size exceeds "
                    "DATAPAW_XLSX_MAX_UNCOMPRESSED_MB"
                )
            if unpacked > max_ratio * max(1, packed):
                raise ValueError(
                    "xlsx compression ratio exceeds "
                    "DATAPAW_XLSX_MAX_COMPRESSION_RATIO"
                )
    except zipfile.BadZipFile as exc:
        raise ValueError("invalid xlsx archive") from exc


def _load_sheet_rows(wb: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = iter(ws.iter_rows(values_only=True))
    try:
        first_row = next(rows)
    except StopIteration:
        return []
    headers = [str(c).strip() if c is not None else "" for c in first_row]
    out: list[dict[str, Any]] = []
    max_rows = _positive_env_int("DATAPAW_XLSX_MAX_ROWS_PER_SHEET", 200000)
    max_cells = _positive_env_int("DATAPAW_XLSX_MAX_CELLS_PER_SHEET", 2000000)
    cell_count = len(first_row)
    for r in rows:
        if len(out) >= max_rows:
            raise ValueError(
                f"sheet {sheet_name!r} exceeds DATAPAW_XLSX_MAX_ROWS_PER_SHEET"
            )
        cell_count += len(r)
        if cell_count > max_cells:
            raise ValueError(
                f"sheet {sheet_name!r} exceeds DATAPAW_XLSX_MAX_CELLS_PER_SHEET"
            )
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        rec: dict[str, Any] = {}
        for i, v in enumerate(r):
            if i >= len(headers) or not headers[i]:
                continue
            rec[headers[i]] = v
        out.append(rec)
    return out


def _coerce_row(sheet: str, raw: dict[str, Any]) -> dict[str, Any]:
    """把 xlsx 原始 cell 值规整成内部表示（多值 split、布尔还原）。

    注意：``enums`` 和 ``column_enums_description`` 存的是 JSON 而非 ``$$$``
    拼接字符串，必须在 ``MULTIVALUE_COLUMNS`` 分支之前单独处理。
    """
    out: dict[str, Any] = {}
    for col in SHEET_COLUMNS[sheet]:
        val = raw.get(col)
        if col == "enums":
            s = str(val or "").strip()
            if s:
                try:
                    out[col] = json.loads(s)
                except (ValueError, TypeError):
                    out[col] = [{"value": p.strip(), "occur_cnt": 0}
                                for p in _split_list(s)]
            else:
                out[col] = []
        elif col == "column_enums_description":
            s = str(val or "").strip()
            if s:
                try:
                    out[col] = json.loads(s)
                except (ValueError, TypeError):
                    out[col] = {}
            else:
                out[col] = {}
        elif col in MULTIVALUE_COLUMNS:
            out[col] = _split_list(val)
        elif col in BOOL_COLUMNS:
            out[col] = _cell_to_bool(val)
        elif col == "depth":
            try:
                out[col] = int(val) if val is not None else 0
            except (TypeError, ValueError):
                out[col] = 0
        else:
            out[col] = str(val).strip() if val is not None else ""
    out["_row_key"] = _row_key(sheet, out)
    return out


def parse_import_workbook(path: Path) -> ParsedImport:
    """解析用户编辑后的 xlsx → ParsedImport（F）。"""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl required for import parsing") from exc

    _validate_xlsx_archive(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        rows_by_sheet: dict[str, list[dict[str, Any]]] = {}
        for sheet in SHEET_ORDER:
            raw_rows = _load_sheet_rows(wb, sheet)
            rows_by_sheet[sheet] = [_coerce_row(sheet, r) for r in raw_rows]
        meta_rows = _load_sheet_rows(wb, "_meta")
    finally:
        wb.close()

    # _meta
    meta: Optional[ExportBaseline] = None
    if meta_rows:
        m = {str(r.get("key") or ""): r.get("value") for r in meta_rows}
        ek_raw = str(m.get("entity_keys") or "")
        try:
            entity_keys = json.loads(ek_raw) if ek_raw else {}
        except (ValueError, TypeError):
            entity_keys = {}
        try:
            exported_at = float(m.get("exported_at") or 0)
        except (TypeError, ValueError):
            exported_at = 0.0
        meta = ExportBaseline(
            export_id=str(m.get("export_id") or ""),
            datasource_id=str(m.get("datasource_id") or ""),
            exported_at=exported_at,
            graph_digest=str(m.get("graph_digest") or ""),
            entity_keys=entity_keys,
        )

    ds_code = ""
    ds_rows = rows_by_sheet.get("datasource") or []
    if ds_rows:
        ds_code = str(ds_rows[0].get("datasource_id") or "").strip()
    elif meta:
        ds_code = meta.datasource_id

    file_keys: dict[str, list[str]] = {
        sheet: sorted({_row_key(sheet, r) for r in rows_by_sheet.get(sheet, [])})
        for sheet in SHEET_ORDER
    }
    return ParsedImport(
        datasource_id=ds_code,
        rows_by_sheet=rows_by_sheet,
        meta=meta,
        file_keys=file_keys,
    )


# ---------------------------------------------------------------------- #
# xlsx → SemanticPayload JSON 转换（from-excel 导入用）
# ---------------------------------------------------------------------- #

def workbook_to_semantic_payload(
    parsed: ParsedImport,
    *,
    target_datasource_id: str = "",
) -> dict[str, Any]:
    """将 ParsedImport（8 sheet bundle）转换为 SemanticImportRequest 兼容的 JSON。

    ``target_datasource_id`` 非空时，替换所有行中的 datasource_id，
    实现「一键切换数据源」。

    返回结构::

        {
            "datasource_id": "...",
            "original_datasource_id": "...",   # 替换前的原始值（便于审计）
            "semantic": { "domains": [...] },    # SemanticPayload
        }
    """
    rows = parsed.rows_by_sheet
    original_ds = parsed.datasource_id
    ds = target_datasource_id.strip() if target_datasource_id else original_ds

    # ---- 按 domain 分组 ----
    domain_rows = rows.get("biz_domain", [])
    domain_names: list[str] = []
    domain_meta: dict[str, dict[str, Any]] = {}
    for dr in domain_rows:
        name = str(dr.get("domain_name") or "").strip()
        if not name or name in domain_meta:
            continue
        domain_names.append(name)
        domain_meta[name] = dr

    # ---- datasets (grouped by domain) ----
    ds_rows = rows.get("dataset", [])
    col_rows = rows.get("dataset_column", [])
    ds_dim_rows = rows.get("dataset_dimension", [])

    # index: (domain, dataset) → [col_rows]
    col_idx: dict[tuple[str, str], list[dict]] = {}
    for c in col_rows:
        key = (str(c.get("domain_name") or ""), str(c.get("dataset_name") or ""))
        col_idx.setdefault(key, []).append(c)

    # index: (domain, dataset, dimension) → ds_dim row
    ds_dim_idx: dict[tuple[str, str], list[dict]] = {}
    for dd in ds_dim_rows:
        key = (str(dd.get("domain_name") or ""), str(dd.get("dimension_name") or ""))
        ds_dim_idx.setdefault(key, []).append(dd)

    # ---- dimensions (grouped by domain) ----
    dim_rows = rows.get("dimension", [])

    # ---- metrics + formulas (grouped by domain) ----
    met_rows = rows.get("metric", [])
    formula_rows = rows.get("metric_formula", [])

    # index: (domain, metric) → [formula rows]
    formula_idx: dict[tuple[str, str], list[dict]] = {}
    for fr in formula_rows:
        key = (str(fr.get("domain_name") or ""), str(fr.get("metric_name") or ""))
        formula_idx.setdefault(key, []).append(fr)

    # ---- build payload ----
    domains_payload: list[dict[str, Any]] = []
    for dname in domain_names:
        dmeta = domain_meta[dname]

        # datasets
        datasets_out: list[dict[str, Any]] = []
        for dsr in ds_rows:
            if str(dsr.get("domain_name") or "") != dname:
                continue
            dsname = str(dsr.get("dataset_name") or "")
            cols_out: list[dict[str, Any]] = []
            for cr in col_idx.get((dname, dsname), []):
                enums_val = cr.get("column_enums")
                enums_list = enums_val if isinstance(enums_val, list) else _split_list(enums_val)
                ced = cr.get("column_enums_description")
                ced_list: list[str] = []
                if isinstance(ced, dict):
                    ced_list = [str(ced.get(e, "")) for e in enums_list]
                elif isinstance(ced, list):
                    ced_list = [str(x) for x in ced]
                samples_val = cr.get("samples")
                samples_list = samples_val if isinstance(samples_val, list) else _split_list(samples_val)
                cols_out.append({
                    "name": str(cr.get("column_name") or ""),
                    "data_type": str(cr.get("data_type") or "text"),
                    "is_primary": cr.get("is_primary") if isinstance(cr.get("is_primary"), bool) else _cell_to_bool(cr.get("is_primary")),
                    "is_nullable": cr.get("is_nullable") if isinstance(cr.get("is_nullable"), bool) else _cell_to_bool(cr.get("is_nullable")),
                    "comment": str(cr.get("column_comment") or ""),
                    "name_cn": str(cr.get("column_name_cn") or ""),
                    "column_type": str(cr.get("column_type") or ""),
                    "enums": enums_list or None,
                    "enums_description": ced_list or None,
                    "samples": samples_list or None,
                    "dimension_type": str(cr.get("dimension_type") or ""),
                })
            parents_raw = str(dsr.get("parents") or "")
            parents_list = [p.strip() for p in parents_raw.split(",") if p.strip()] if parents_raw else []
            datasets_out.append({
                "name": dsname,
                "description": str(dsr.get("dataset_comment") or ""),
                "dataset_type": str(dsr.get("dataset_type") or "OLAP"),
                "sql": str(dsr.get("sql_content") or "*"),
                "parents": parents_list,
                "columns": cols_out,
            })

        # dimensions
        dims_out: list[dict[str, Any]] = []
        for dr in dim_rows:
            if str(dr.get("domain_name") or "") != dname:
                continue
            dim_name = str(dr.get("dimension_name") or "")
            synonyms_val = dr.get("synonyms")
            aliases = synonyms_val if isinstance(synonyms_val, list) else _split_list(synonyms_val)
            enums_raw = dr.get("enums") or []
            values_out: list[dict[str, Any]] = []
            if isinstance(enums_raw, list):
                for ev in enums_raw:
                    if isinstance(ev, dict):
                        values_out.append({"value": str(ev.get("value", "")), "aliases": []})
                    else:
                        values_out.append({"value": str(ev), "aliases": []})

            bindings_out: list[dict[str, Any]] = []
            for dd in ds_dim_idx.get((dname, dim_name), []):
                expr = str(dd.get("calculate_expr") or "")
                bindings_out.append({
                    "dataset": str(dd.get("dataset_name") or ""),
                    "calculate_expr": expr,
                    "binding_type": str(dd.get("dimension_type") or "OLAP维度"),
                    "data_type": str(dd.get("data_type") or "text"),
                    "aliases": [],
                })

            dims_out.append({
                "name": dim_name,
                "description": str(dr.get("description") or ""),
                "aliases": aliases,
                "parent_dimension": str(dr.get("parent_name") or ""),
                "hierarchy_level": int(dr.get("depth") or 0) if not isinstance(dr.get("depth"), int) else dr.get("depth", 0),
                "is_display_dimension": dr.get("is_visible") if isinstance(dr.get("is_visible"), bool) else _cell_to_bool(dr.get("is_visible")),
                "is_contribution_dimension": dr.get("is_attribution") if isinstance(dr.get("is_attribution"), bool) else _cell_to_bool(dr.get("is_attribution")),
                "bindings": bindings_out,
                "values": values_out,
            })

        # metrics
        metrics_out: list[dict[str, Any]] = []
        for mr in met_rows:
            if str(mr.get("domain_name") or "") != dname:
                continue
            mname = str(mr.get("metric_name") or "")
            syn_val = mr.get("synonyms")
            m_aliases = syn_val if isinstance(syn_val, list) else _split_list(syn_val)
            tag_val = mr.get("tags")
            m_tags = tag_val if isinstance(tag_val, list) else _split_list(tag_val)

            formulas_out: list[dict[str, Any]] = []
            derived_from_out: list[dict[str, str]] = []
            for fr in formula_idx.get((dname, mname), []):
                formulas_out.append({
                    "dataset": str(fr.get("dataset_name") or ""),
                    "formula": str(fr.get("formula") or ""),
                    "formula_evidence": str(fr.get("formula_evidence") or ""),
                    "date_range": str(fr.get("date_range") or ""),
                    "is_primary": len(formulas_out) == 0,
                })
                df_val = fr.get("derived_from")
                df_list = df_val if isinstance(df_val, list) else _split_list(df_val)
                for d in df_list:
                    d = str(d).strip()
                    if not d:
                        continue
                    if ":" in d:
                        name_part, role_part = d.split(":", 1)
                        derived_from_out.append({"metric_name": name_part.strip(), "role": role_part.strip()})
                    else:
                        derived_from_out.append({"metric_name": d, "relation_type": "ratio_decompose"})

            analyzed_by: list[str] = []
            for dd in ds_dim_rows:
                if str(dd.get("domain_name") or "") == dname:
                    dim_n = str(dd.get("dimension_name") or "")
                    if dim_n and dim_n not in analyzed_by:
                        analyzed_by.append(dim_n)

            metrics_out.append({
                "name": mname,
                "description": str(mr.get("description") or ""),
                "unit": str(mr.get("unit") or ""),
                "is_north_star": mr.get("is_polaris") if isinstance(mr.get("is_polaris"), bool) else _cell_to_bool(mr.get("is_polaris")),
                "is_display_distribution": mr.get("show_distribution") if isinstance(mr.get("show_distribution"), bool) else _cell_to_bool(mr.get("show_distribution")),
                "is_display": mr.get("is_visible") if isinstance(mr.get("is_visible"), bool) else _cell_to_bool(mr.get("is_visible")),
                "aliases": m_aliases,
                "tags": m_tags,
                "formulas": formulas_out,
                "analyzed_by": [],
                "derived_from": derived_from_out,
            })

        d_aliases = dmeta.get("aliases")
        if isinstance(d_aliases, str):
            d_aliases = _split_list(d_aliases)
        elif not isinstance(d_aliases, list):
            d_aliases = []

        domains_payload.append({
            "name": dname,
            "display_name": str(dmeta.get("display_name") or dname),
            "description": str(dmeta.get("description") or ""),
            "aliases": d_aliases,
            "datasets": datasets_out,
            "dimensions": dims_out,
            "metrics": metrics_out,
        })

    result: dict[str, Any] = {
        "datasource_id": ds,
        "semantic": {"domains": domains_payload},
    }
    if ds != original_ds:
        result["original_datasource_id"] = original_ds
    return result


__all__ = [
    "ExportBaseline",
    "ParsedImport",
    "SHEET_COLUMNS",
    "SHEET_KEY_COLUMNS",
    "MULTIVALUE_COLUMNS",
    "SHEET_ORDER",
    "build_bundle_from_graph",
    "export_to_workbook",
    "save_workbook",
    "parse_import_workbook",
    "workbook_to_semantic_payload",
]
