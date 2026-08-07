from __future__ import annotations

import io

import aiosqlite
from openpyxl import load_workbook

from semantic_config.db import now_iso
from semantic_config.errors import BadRequestError

MAX_BYTES = 50 * 1024 * 1024

# 已知 sheet 与拓扑顺序（未知 sheet 跳过，如 _meta）
KNOWN_SHEETS = [
    "datasource", "biz_domain", "dataset", "dataset_column",
    "dimension", "dataset_dimension", "metric", "metric_formula",
]

SUMMARY_KEY = {
    "datasource": "datasource", "biz_domain": "biz_domain", "dataset": "dataset",
    "dataset_column": "dataset_column", "dimension": "dimension",
    "dataset_dimension": "dataset_dimension", "metric": "metric", "metric_formula": "metric_formula",
}


class _RowError(Exception):
    def __init__(self, sheet: str, row: int, message: str):
        self.sheet = sheet
        self.row = row
        self.message = message


def _norm(v) -> str | None:
    """空单元格 / '\\N' -> None；数值转字符串；其余去空白。"""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return None if s in ("", "\\N") else s
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _to_int(v) -> int | None:
    s = _norm(v)
    if s is None:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _to_bool_int(v, default: int) -> int:
    """真布尔列：1->1、0->0；空 -> 默认值。"""
    s = _norm(v)
    if s is None:
        return default
    return 1 if s in ("1", "true", "True", "TRUE") else 0


def _to_yn(v, default: str) -> str:
    """is_primary/is_nullable：Excel 0/1 -> 'N'/'Y'；空 -> 默认。"""
    s = _norm(v)
    if s is None:
        return default
    if s in ("1", "Y", "y"):
        return "Y"
    if s in ("0", "N", "n"):
        return "N"
    return default


def _read_sheets(content: bytes) -> dict[str, list[dict]]:
    """解析 xlsx -> {sheet_name: [row_dict, ...]}（row_dict 含 _row=excel行号）。仅保留已知 sheet。"""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise BadRequestError("解析失败 / 不是有效 xlsx") from e

    result: dict[str, list[dict]] = {}
    for ws in wb.worksheets:
        if ws.title not in KNOWN_SHEETS:
            continue
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue
        headers = [(_norm(h) or "") for h in header]
        data = []
        for i, raw in enumerate(rows_iter, start=2):
            if raw is None or all(c is None for c in raw):
                continue
            row = {headers[j]: raw[j] if j < len(raw) else None for j in range(len(headers))}
            row["_row"] = i
            data.append(row)
        result[ws.title] = data
    wb.close()
    return result


# ---------------- 引用解析（查库，能看到本事务内已 upsert 的记录） ----------------
async def _find_id(db, sql: str, params: tuple) -> int | None:
    async with db.execute(sql, params) as cur:
        row = await cur.fetchone()
    return int(row["id"]) if row else None


async def _resolve_ds(db, datasource_id: str | None, sheet: str, rn: int) -> str:
    """校验数据源编码存在，返回该编码（各子表 datasource_id 直接存编码）。"""
    if not datasource_id:
        raise _RowError(sheet, rn, "datasource_id 不能为空")
    found = await _find_id(db, "SELECT id FROM datasource WHERE datasource_id = ? AND is_deleted = 0", (datasource_id,))
    if found is None:
        raise _RowError(sheet, rn, f"数据源不存在: datasource_id={datasource_id}")
    return datasource_id


async def _resolve_domain(db, ds_id: str, name: str | None, sheet: str, rn: int) -> int:
    if not name:
        raise _RowError(sheet, rn, "domain_name 不能为空")
    dom_id = await _find_id(
        db, "SELECT id FROM biz_domain WHERE datasource_id = ? AND domain_name = ? AND is_deleted = 0",
        (ds_id, name),
    )
    if dom_id is None:
        raise _RowError(sheet, rn, f"业务域不存在: domain_name={name}")
    return dom_id


async def _resolve_dataset(db, domain_id: int, name: str | None, sheet: str, rn: int) -> int:
    if not name:
        raise _RowError(sheet, rn, "dataset_name 不能为空")
    dsid = await _find_id(
        db, "SELECT id FROM dataset_meta WHERE domain_id = ? AND dataset_name = ? AND is_deleted = 0",
        (domain_id, name),
    )
    if dsid is None:
        raise _RowError(sheet, rn, f"数据集不存在: dataset_name={name}")
    return dsid


async def _resolve_dimension(db, domain_id: int, name: str | None, sheet: str, rn: int) -> int:
    if not name:
        raise _RowError(sheet, rn, "dimension_name 不能为空")
    did = await _find_id(
        db, "SELECT id FROM dimension WHERE domain_id = ? AND dimension_name = ? AND is_deleted = 0",
        (domain_id, name),
    )
    if did is None:
        raise _RowError(sheet, rn, f"维度不存在: dimension_name={name}")
    return did


async def _resolve_metric(db, domain_id: int, name: str | None, sheet: str, rn: int) -> int:
    if not name:
        raise _RowError(sheet, rn, "metric_name 不能为空")
    mid = await _find_id(
        db, "SELECT id FROM metric_lib WHERE domain_id = ? AND metric_name = ? AND is_deleted = 0",
        (domain_id, name),
    )
    if mid is None:
        raise _RowError(sheet, rn, f"指标不存在: metric_name={name}")
    return mid


# ---------------- 通用 upsert ----------------
async def _upsert(db, table: str, nk: dict, cols: dict) -> int:
    where = " AND ".join(f"{k} IS ?" for k in nk)
    found = await _find_id(db, f"SELECT id FROM {table} WHERE {where} AND is_deleted = 0", tuple(nk.values()))
    ts = now_iso()
    if found is not None:
        sets = ", ".join(f"{k} = ?" for k in cols) + ", updated_at = ?"
        await db.execute(
            f"UPDATE {table} SET {sets} WHERE id = ?", (*cols.values(), ts, found)
        )
        return found
    keys = ", ".join([*cols.keys(), "created_at", "updated_at"])
    ph = ", ".join(["?"] * (len(cols) + 2))
    async with db.execute(
        f"INSERT INTO {table} ({keys}) VALUES ({ph}) RETURNING id", (*cols.values(), ts, ts)
    ) as cur:
        row = await cur.fetchone()
    return int(row["id"])


# ---------------- 各 sheet 处理 ----------------
async def _do_datasource(db, r) -> None:
    datasource_id = _norm(r.get("datasource_id"))
    if not datasource_id:
        raise _RowError("datasource", r["_row"], "datasource_id 不能为空")
    await _upsert(db, "datasource", {"datasource_id": datasource_id}, {
        "datasource_id": datasource_id,
        "datasource_name": _norm(r.get("datasource_name")),
        "datasource_type": _norm(r.get("datasource_type")),
    })


async def _do_biz_domain(db, r) -> None:
    rn = r["_row"]
    ds_id = await _resolve_ds(db, _norm(r.get("datasource_id")), "biz_domain", rn)
    name = _norm(r.get("domain_name"))
    if not name:
        raise _RowError("biz_domain", rn, "domain_name 不能为空")
    await _upsert(db, "biz_domain", {"datasource_id": ds_id, "domain_name": name}, {
        "datasource_id": ds_id, "domain_name": name,
        "display_name": _norm(r.get("display_name")),
        "description": _norm(r.get("description")),
        "aliases": _norm(r.get("aliases")),
    })


async def _do_dataset(db, r) -> None:
    rn = r["_row"]
    ds_id = await _resolve_ds(db, _norm(r.get("datasource_id")), "dataset", rn)
    dom_id = await _resolve_domain(db, ds_id, _norm(r.get("domain_name")), "dataset", rn)
    name = _norm(r.get("dataset_name"))
    if not name:
        raise _RowError("dataset", rn, "dataset_name 不能为空")
    await _upsert(db, "dataset_meta", {"domain_id": dom_id, "dataset_name": name}, {
        "datasource_id": ds_id, "domain_id": dom_id, "dataset_name": name,
        "dataset_comment": _norm(r.get("dataset_comment")),
        "dataset_type": _norm(r.get("dataset_type")),
        "sql_content": _norm(r.get("sql_content")),
        "parents": _norm(r.get("parents")),
    })


async def _do_dataset_column(db, r) -> None:
    rn = r["_row"]
    ds_id = await _resolve_ds(db, _norm(r.get("datasource_id")), "dataset_column", rn)
    dom_id = await _resolve_domain(db, ds_id, _norm(r.get("domain_name")), "dataset_column", rn)
    dataset_id = await _resolve_dataset(db, dom_id, _norm(r.get("dataset_name")), "dataset_column", rn)
    col = _norm(r.get("column_name"))
    if not col:
        raise _RowError("dataset_column", rn, "column_name 不能为空")
    await _upsert(db, "dataset_column_meta", {"dataset_id": dataset_id, "column_name": col}, {
        "dataset_id": dataset_id, "datasource_id": ds_id, "domain_id": dom_id, "column_name": col,
        "column_name_cn": _norm(r.get("column_name_cn")),
        "data_type": _norm(r.get("data_type")),
        "column_type": _norm(r.get("column_type")),
        "dimension_type": _norm(r.get("dimension_type")),
        "column_comment": _norm(r.get("column_comment")),
        "column_enums": _norm(r.get("column_enums")),
        "column_enums_description": _norm(r.get("column_enums_description")),
        "samples": _norm(r.get("samples")),
        "is_primary": _to_yn(r.get("is_primary"), "N"),
        "is_nullable": _to_yn(r.get("is_nullable"), "Y"),
    })


async def _do_dimension(db, r) -> None:
    rn = r["_row"]
    ds_id = await _resolve_ds(db, _norm(r.get("datasource_id")), "dimension", rn)
    dom_id = await _resolve_domain(db, ds_id, _norm(r.get("domain_name")), "dimension", rn)
    name = _norm(r.get("dimension_name"))
    if not name:
        raise _RowError("dimension", rn, "dimension_name 不能为空")
    await _upsert(db, "dimension", {"domain_id": dom_id, "dimension_name": name}, {
        "datasource_id": ds_id, "domain_id": dom_id, "dimension_name": name,
        "description": _norm(r.get("description")),
        "parent_name": _norm(r.get("parent_name")),
        "depth": _to_int(r.get("depth")),
        "synonyms": _norm(r.get("synonyms")),
        "is_visible": _to_bool_int(r.get("is_visible"), 1),
        "is_attribution": _to_bool_int(r.get("is_attribution"), 1),
        "enums": _norm(r.get("enums")),
    })


async def _do_dataset_dimension(db, r) -> None:
    rn = r["_row"]
    ds_id = await _resolve_ds(db, _norm(r.get("datasource_id")), "dataset_dimension", rn)
    dom_id = await _resolve_domain(db, ds_id, _norm(r.get("domain_name")), "dataset_dimension", rn)
    dataset_id = await _resolve_dataset(db, dom_id, _norm(r.get("dataset_name")), "dataset_dimension", rn)
    dim_id = await _resolve_dimension(db, dom_id, _norm(r.get("dimension_name")), "dataset_dimension", rn)
    await _upsert(db, "dataset_dimension", {"dataset_id": dataset_id, "dimension_id": dim_id}, {
        "dataset_id": dataset_id, "dimension_id": dim_id, "datasource_id": ds_id, "domain_id": dom_id,
        "calculate_expr": _norm(r.get("calculate_expr")),
        "dimension_type": _norm(r.get("dimension_type")),
        "data_type": _norm(r.get("data_type")),
    })


async def _do_metric(db, r) -> None:
    rn = r["_row"]
    ds_id = await _resolve_ds(db, _norm(r.get("datasource_id")), "metric", rn)
    dom_id = await _resolve_domain(db, ds_id, _norm(r.get("domain_name")), "metric", rn)
    name = _norm(r.get("metric_name"))
    if not name:
        raise _RowError("metric", rn, "metric_name 不能为空")
    await _upsert(db, "metric_lib", {"domain_id": dom_id, "metric_name": name}, {
        "datasource_id": ds_id, "domain_id": dom_id, "metric_name": name,
        "description": _norm(r.get("description")),
        "unit": _norm(r.get("unit")),
        "is_polaris": _to_bool_int(r.get("is_polaris"), 0),
        "show_distribution": _to_bool_int(r.get("show_distribution"), 0),
        "is_visible": _to_bool_int(r.get("is_visible"), 1),
        "synonyms": _norm(r.get("synonyms")),
        "tags": _norm(r.get("tags")),
    })


async def _do_metric_formula(db, r) -> None:
    rn = r["_row"]
    ds_id = await _resolve_ds(db, _norm(r.get("datasource_id")), "metric_formula", rn)
    dom_id = await _resolve_domain(db, ds_id, _norm(r.get("domain_name")), "metric_formula", rn)
    metric_id = await _resolve_metric(db, dom_id, _norm(r.get("metric_name")), "metric_formula", rn)
    dataset_id = await _resolve_dataset(db, dom_id, _norm(r.get("dataset_name")), "metric_formula", rn)
    date_range = _norm(r.get("date_range"))
    await _upsert(
        db, "metric_formula_lib",
        {"metric_id": metric_id, "dataset_id": dataset_id, "date_range": date_range},
        {
            "metric_id": metric_id, "dataset_id": dataset_id, "datasource_id": ds_id, "domain_id": dom_id,
            "formula": _norm(r.get("formula")),
            "date_range": date_range,
            "formula_evidence": _norm(r.get("formula_evidence")),
            "derived_from": _norm(r.get("derived_from")),
            "evidence_ext": _norm(r.get("evidence_ext")),
        },
    )


_HANDLERS = {
    "datasource": _do_datasource,
    "biz_domain": _do_biz_domain,
    "dataset": _do_dataset,
    "dataset_column": _do_dataset_column,
    "dimension": _do_dimension,
    "dataset_dimension": _do_dataset_dimension,
    "metric": _do_metric,
    "metric_formula": _do_metric_formula,
}


async def import_excel(db: aiosqlite.Connection, filename: str, content: bytes) -> dict:
    # 文件级校验
    if not content:
        raise BadRequestError("上传文件不能为空")
    if not (filename or "").lower().endswith(".xlsx"):
        raise BadRequestError("仅支持 .xlsx 格式")
    if len(content) > MAX_BYTES:
        raise BadRequestError("文件大小不能超过 50MB")

    sheets = _read_sheets(content)
    if not sheets:
        raise BadRequestError("未找到可导入的 sheet（datasource/biz_domain/...）")

    # 单事务：按拓扑序逐行 upsert，任一行错误整体回滚
    summary: dict[str, int] = {}
    try:
        await db.execute("BEGIN")
        for sheet in KNOWN_SHEETS:
            rows = sheets.get(sheet)
            if not rows:
                continue
            handler = _HANDLERS[sheet]
            cnt = 0
            for r in rows:
                await handler(db, r)
                cnt += 1
            summary[SUMMARY_KEY[sheet]] = cnt
        await db.commit()
        return {"success": True, "summary": summary, "errors": []}
    except _RowError as e:
        await db.rollback()
        return {
            "success": False,
            "summary": {},
            "errors": [{"sheet": e.sheet, "row": e.row, "message": e.message}],
        }
    except Exception:
        await db.rollback()
        raise
